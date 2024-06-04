import openpyxl
from openpyxl.styles import PatternFill

def highlight_rows(file_path):
    # Load the workbook and select the active worksheet
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # Define the yellow fill color
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Iterate through the rows in the worksheet
    for row in ws.iter_rows(min_row=2):  # Assuming the first row is a header
        highlight = False
        for cell in row[6:12]:  # Columns G to L (0-indexed: G is 6, L is 11)
            if cell.value is not None and isinstance(cell.value, (int, float)) and cell.value <= 3:
                highlight = True
                break

        # If any cell in the range G to L is 3 or below, highlight the entire row
        if highlight:
            for cell in row:
                cell.fill = yellow_fill

    # Save the modified workbook
    wb.save(file_path)

# Use the function
def emain(file):
  highlight_rows(file)
  return file

  